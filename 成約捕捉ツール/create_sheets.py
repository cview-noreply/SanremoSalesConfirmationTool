# -*- coding: utf-8 -*-
"""
Project: サンレモ成約捕捉
File: creat_sheets.py
Description: 
    1. 作業フォルダの作成
    2. 案件管理シートの作成
    3. 作成した案件管理シートのチェック(check_sheets.py 内のチェックルールを援用)、チェック結果の作成
    4. 送信対象クライアント一覧の作成
    5. 案件管理シートの送信実績取込用ファイル作成

Copyright (c) 2026 SCSK ServiceWare Corporation.
All rights reserved.
"""

import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import shutil
from pathlib import Path
import numpy as np

from check_sheets import SheetChecker
from utils import (
    # 設定値
    set_config,
    # 独自エラー
    AppError, ZeroDataError, NotSelectError, ReferencePathError, NoSheetError,
    # 共通関数
    get_pw, select_folder, select_file, search_file, selectfile_to_df, serial_filepath, sanitize_filename, get_base_dir,
    create_name, create_filename, exist_folders
)


# ==== 日付文字列生成 ====
YYYYMMDD = datetime.date.today().strftime("%Y%m%d")
YYYYMM = datetime.date.today().strftime("%Y%m")

# ==== 外部参照 ====
config = set_config()


# ============================================
# フォルダーの作成
# ============================================
ROOT_FOLDER = Path(config['ルートフォルダ'])
BASE_FOLDER = ROOT_FOLDER / '02_案件管理シート作成' 
WORKING_FOLDER = BASE_FOLDER / YYYYMM
CSV_FOLDER = WORKING_FOLDER / '01_CSV'
SHEETS_FOLDER = WORKING_FOLDER / '02_作成済み案件管理シート'
RESULT_FOLDER = WORKING_FOLDER / '99_処理結果'

# フォルダの作成（既存のものは触れない）
def make_folders():
    WORKING_FOLDER.mkdir(parents=True, exist_ok=True)
    CSV_FOLDER.mkdir(exist_ok=True)
    SHEETS_FOLDER.mkdir(exist_ok=True)
    RESULT_FOLDER.mkdir(exist_ok=True)

# ============================================
# 案件管理シートの作成
# ============================================
def create_sheets():

    exist_folders([BASE_FOLDER, WORKING_FOLDER, CSV_FOLDER, SHEETS_FOLDER, RESULT_FOLDER])

    result = []

    template_filepath = config['ファイルパス']['案件管理シートFMT']
    target_sheet_name = config['案件管理シート']['シート名']
    restart_interval = int(config['xw_RESTART_INTERVAL'])

    df = selectfile_to_df('案件管理シート全件リスト.csv', WORKING_FOLDER)
    if len(df) == 0:
        raise ZeroDataError('データが0件です')

    group = df.groupby('振分先コード')
    group_list = list(group)

    app = None
    wb = None
    ws = None

    def _boot_app_and_open_template():
        """Excelアプリを起動してテンプレートを開き (app, wb, ws) を返す"""
        _app = xw.App(visible=False)
        _app.screen_updating = False
        _app.display_alerts = False
        try:
            _wb = _app.books.open(template_filepath)
            _ws = _wb.sheets[target_sheet_name]
        except Exception:
            _app.quit()
            raise RuntimeError(f'テンプレートまたはシート({target_sheet_name})が見つかりません')
        return _app, _wb, _ws

    err_count = 0
    ok_count = 0
    exist_count = 0

    try:
        for i, (code, group_df) in enumerate(group_list):

            # ── Excelアプリの起動 or 定期再起動 ──────────────────────────
            # ループの先頭で判断し、wb/ws も同時に取得
            if app is None or (i > 0 and i % restart_interval == 0):
                if app is not None:
                    try:
                        wb.close()   # ★ app.quit() 前に必ず wb を閉じる
                    except Exception:
                        pass
                    app.quit()

                app, wb, ws = _boot_app_and_open_template()

            # ── 各種情報取得 ──────────────────────────────────────────────
            kigyo_nm         = group_df['企業名'].iloc[0]
            furiwakesaki_nm  = group_df['振分先名'].iloc[0]
            rows             = len(group_df)

            print(f'[{i+1}/{len(group_list)}] 処理開始: {code}|{kigyo_nm}|{furiwakesaki_nm}')

            entry = {
                '振分先コード': code,
                '企業名':       kigyo_nm,
                '振分先名':     furiwakesaki_nm,
                'pw':           '',
                'データ件数':   rows,
                '作成日時':     f'{datetime.datetime.now():%Y/%m/%d %H:%M:%S}',
                'ファイル名':   '',
                'パス内禁止文字': ''
            }

            # ── PW取得 ───────────────────────────────────────────────────
            pw = get_pw(code)
            if not pw:
                entry['pw'] = '一覧内になし'
                result.append(entry)
                print(' ⚠️ pwなし')
                err_count += 1
                continue
            entry['pw'] = pw

            # ── 保存ファイル名作成 & 禁止文字チェック ────────────────────
            save_filename = create_filename(code, kigyo_nm, furiwakesaki_nm)
            rep_filename  = sanitize_filename(save_filename)
            if rep_filename != save_filename:
                save_filename = rep_filename
                entry['パス内禁止文字'] = 'あり'
            entry['ファイル名'] = save_filename

            # ── 作成済みスキップ ─────────────────────────────────────────
            save_filepath = SHEETS_FOLDER / save_filename
            if save_filepath.exists():
                entry['作成日時'] = '過去に作成済み'
                result.append(entry)
                print(' ⚠️ 作成済み')
                exist_count += 1
                continue

            # ── Excel書き込み ────────────────────────────────────────────
            ws.range('D2').value = code
            ws.range('D3').value = create_name(kigyo_nm, furiwakesaki_nm)

            ws.range('F:F').number_format = '@'
            ws.range('U:U').number_format = '@'
            ws.range('B9').options(index=False, header=False).value = group_df.iloc[:, 3:]

            last_row = ws.used_range.last_cell.row

            # 空列(メモ欄)挿入
            start_col, end_col = config['案件管理シート']['空列範囲'].split(':')
            ws.range(f'{start_col}9:{end_col}{last_row}').insert(shift='right')
            
            # 空列挿入でズレた分の右端を削除
            last_col = config['案件管理シート']['全体範囲'].split(':')[1]
            del_col  = get_column_letter(column_index_from_string(last_col) + 1)
            ws.range(f'{del_col}9:AZ{last_row}').delete(shift='left')

            # CL入力列のセルロックを解除
            editable_left, editable_right = config['案件管理シート']['入力範囲'].split(':')
            ws.range(f'{editable_left}9:{editable_right}{9 + rows - 1}').api.Locked = False

            ws.api.Protect(
                Password=config['案件管理シート']['保護パスワード'],
                AllowFiltering=True,
                AllowFormattingColumns=True,
                AllowFormattingRows=True
            )

            # SaveAs でファイルを書き出した後、wb は「別名で開いている状態」になる。
            # 次ループでテンプレートとして再利用するため SaveAs 後に閉じ、
            # テンプレートを再度開いて wb/ws を更新する。
            wb.api.SaveAs(str(save_filepath), Password=pw)
            wb.close()
            wb = app.books.open(template_filepath)
            ws = wb.sheets[target_sheet_name]

            result.append(entry)
            print(' ✔  完了')
            ok_count += 1

    finally:
        if app is not None:
            try:
                wb.close()
            except Exception:
                pass
            app.quit()

    print(f' ✅ 作成完了: {ok_count} | 作成済み: {exist_count} | 作成エラー: {err_count} | 合計: {len(group_list)}')

    result_df = pd.DataFrame(result)
    result_df.index = range(1, len(result_df) + 1)
    base_name     = f'{YYYYMMDD}_案件管理シート作成結果メモ'
    save_filepath = serial_filepath(RESULT_FOLDER, base_name, '.xlsx')
    result_df.to_excel(save_filepath)

    print(f' >> {save_filepath} を作成')    

# ============================================
# 案件管理シートのチェック
# ============================================
def check_sheets_at_creation():

    exist_folders([BASE_FOLDER])

    dialog_title = f'対象の日付の作業フォルダを選択してください(YYYYMM)'
    print(dialog_title)
    selected_working_folder = select_folder(title=dialog_title, initial_dir=BASE_FOLDER)

    data_df = selectfile_to_df('案件管理シート全件リスト.csv', selected_working_folder)
    if len(data_df) == 0: raise ZeroDataError('データが0件です')
        
    count_dict = data_df['振分先コード'].value_counts().to_dict()
    count_dict_len = len(count_dict)


    target_folder = selected_working_folder / '02_作成済み案件管理シート'
    result_folder = selected_working_folder / '99_処理結果'

    target_files = [f for f in target_folder.glob("*.xlsx") if not f.name.startswith('~$')]
    target_files_len = len(target_files)

    if target_files_len == 0:
        raise ZeroDataError(f'02_作成済み案件管理シートが空です')

    result = []
    ng_count = 0
    ok_count = 0

    for j, file in enumerate(target_files):
        # 進捗
        print(f'[{j+1}/{len(target_files)}] 処理開始: {str(file.name)}')


        # チェック用クラスの作成
        checker = SheetChecker(str(file))
        # 情報の取得
        res = checker.get_info()

        report = {
            '振分先コード': checker.furiwakesaki_code or "",
            '企業名': checker.kigyo_name or "",
            '振分先名': checker.furiwakesaki_name or "",
            'ファイル名': str(file.name),
            'トータルの不備結果': '',
            'PWチェック': '',
            'データ貼り付け、セルロック位置チェック': '',
            'ファイル内反響数チェック': '',
            'ファイル内反響数': '',
            '一覧データ内の反響数': '',
            'ファイル名称チェック': '',
            'CL入力欄文字化けチェック': '',
            'CL入力欄日付データ型チェック': '',
            'CL入力欄文字列型チェック': '',        
        }
        
        match res:
            case 'CodeErr':
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ 振分先コードを取得できません')
                ng_count += 1
                continue
            case 'PwErr':
                report['PWチェック'] = 'NG'
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ パスワードを取得できません')
                ng_count += 1
                continue
            case 'OpenErr':
                report['PWチェック'] = 'NG'
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ ファイルを開けません')
                ng_count += 1
                continue
            case 'DataFrameErr':
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ シート内の情報を取得できません')
                ng_count += 1
                continue
            case 'OpenpyxlErr':
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ シート内の情報を取得できません')
                ng_count += 1
                continue
            case 'NameErr':
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ 社名を取得できません')
                ng_count += 1
                continue
            case None:        
                report['PWチェック'] = 'OK'
    


        report['データ貼り付け、セルロック位置チェック'] = checker.check_cell_protection() and checker.check_input_range()
        report['ファイル内反響数'] = checker.row_count
        report['一覧データ内の反響数'] = count_dict.get(checker.furiwakesaki_code, 0)
        report['ファイル内反響数チェック'] = checker.row_count == count_dict.get(checker.furiwakesaki_code, 0)
        report['ファイル名称チェック'] = checker.check_filename()
        report['CL入力欄文字化けチェック'] = checker.check_mojibake()
        report['CL入力欄日付データ型チェック'] = all([
            checker.check_is_date('①最終更新日​'),
            checker.check_is_date('③初回来場日​'),
            checker.check_is_date('④契約予定日​'),
        ])
        report['CL入力欄文字列型チェック'] = checker.check_is_num()
        
        # トータルで
        check_targets = [
            report['データ貼り付け、セルロック位置チェック'],
            report['ファイル内反響数チェック'],
            report['ファイル名称チェック'],
            report['CL入力欄文字化けチェック'],
            report['CL入力欄日付データ型チェック'],
            report['CL入力欄文字列型チェック']
        ]

        # 全部TrueならTrue、一つでもFalseがあればFalse
        report['トータルの不備結果'] = 'OK' if all(check_targets) else 'NG'

        result.append(report)
        print(f' ✔ 完了')
        ok_count += 1

    print(f' ✅ チェック完了: {ok_count} | チェック不可: {ng_count} | 合計: {len(target_files)}')

    print(f'作成チェック結果を出力します...')
    
    # 結果を出力
    result_df = pd.DataFrame(result)
    result_df = result_df.replace([True, np.True_], 'OK').replace([False, np.False_], 'NG')

    # チェック結果テンプレート
    template_filepath = Path(config['ファイルパス']['チェック結果FMT'])

    # 保存ファイル名(すでにある場合はリネームして新規で作成)
    base_name = f"{YYYYMMDD}_案件管理シート作成チェック結果"
    save_filepath = serial_filepath(result_folder, base_name, '.xlsx')
    shutil.copy(template_filepath, save_filepath)

    # 全体の結果書き込み
    with xw.App(visible=False) as app:
        wb = xw.Book((save_filepath))
        ws = wb.sheets[0]  # 1番左のシートを指定
        
        ws.range("A:A").number_format = '@'
        ws.range('A10').options(index=False, header=False).value = result_df

        ws.range('C3').value = 'OK' if count_dict_len == target_files_len else 'NG'
        ws.range('C4').value = str(target_files_len)
        ws.range('C5').value = str(count_dict_len)
        ws.range('C6').value = 'OK' if all(result_df['トータルの不備結果'] == 'OK') else 'NG'
    
        # 保存
        wb.save()
        wb.close()


    print(f' >> {str(save_filepath)} を作成')


# ============================================
# 送付対象クライアント一覧の作成
# ============================================
def create_send_list():

    exist_folders([BASE_FOLDER])

    title = f'案件管理シート作成チェック結果.xlsxを選択してください'
    print(title)
    result_filepath = select_file(title=title, file_types=[('EXCELファイル', '*.xlsx')], initial_dir=BASE_FOLDER)
    result_folder = result_filepath.parent

    print(f' << {result_filepath} を取得')
        
    # 結果OKで絞り込み
    result_df = pd.read_excel(result_filepath, engine='openpyxl', skiprows=8, usecols='A:E', dtype=str)
    ok_df = result_df[result_df['トータルの不備結果'] == 'OK']
    ok_df = ok_df.iloc[:, :4]   # 4列目まで使用

    # データが0件（！エラー出力）
    if len(ok_df) == 0: raise ZeroDataError('OKが0件です。')


    # == フォーマット読込 ==
    template_filepath = config['ファイルパス']['送付対象クライアント一覧FMT']

    # 保存ファイル名(すでにある場合はリネームして新規で作成)
    base_name = f"{YYYYMMDD}_案件管理シート送付対象クライアント一覧"
    save_filepath = serial_filepath(result_folder, base_name, '.xlsx')
    shutil.copy(template_filepath, save_filepath)

    # == 一括で書きこみ ==
    with xw.App(visible=False) as app:
        wb = xw.Book((save_filepath))
        ws = wb.sheets[0]  # 1番左のシートを指定
        
        ws.range("A:D").number_format = '@'
        ws.range("A2").options(index=False, header=False).value = ok_df
        
        # 保存
        wb.save()
        wb.close()

    print(f' >> {save_filepath} を作成')


# ============================================
# 案件管理シートの送信実績取込用ファイル作成
# ============================================
def create_send_jisseki():

    exist_folders([BASE_FOLDER])

    # 結果取得
    title = f'案件管理シート作成チェック結果.xlsxを選択してください'
    print(title)
    result_filepath = select_file(title=title, file_types=[('EXCELファイル', '*.xlsx')], initial_dir=BASE_FOLDER)
    result_folder = result_filepath.parent

    result_df = pd.read_excel(result_filepath, engine='openpyxl', skiprows=8, usecols='A:E', dtype=str)
    ok_code_list = result_df[result_df['トータルの不備結果'] == 'OK']['振分先コード'].tolist()

    # csvファイルの取得
    sojushin_df = selectfile_to_df("案件管理シート送受信実績一覧.csv", BASE_FOLDER)
    existing_codes = set(sojushin_df['振分先コード'])
    # 日付データの整形
    sojushin_df['案件管理シート送信実施日'] = pd.to_datetime(sojushin_df['案件管理シート送信実施日']).dt.strftime('%Y/%m/%d')
    sojushin_df['案件管理シート回収日'] = pd.to_datetime(sojushin_df['案件管理シート回収日']).dt.strftime('%Y/%m/%d')


    # 今日の日付
    today_str = datetime.date.today().strftime('%Y/%m/%d')
    
    # 追加行の作成
    add_rows = []
    for code in ok_code_list:
        row = {
                'レコードの開始行': '',
                '振分先コード': code,
                '案件管理シート送信実施日': today_str,
                '案件管理シート回収日': '',
                '回収時案件管理シート不備チェック': '',
                '案件管理シート履歴': '',
            }
        
        if not code in existing_codes:
            row['レコードの開始行'] = '*'

        add_rows.append(row)


    add_df = pd.DataFrame(add_rows)
    output_df = pd.concat([sojushin_df, add_df], ignore_index=True)

    # 並べ替え
    output_df = output_df.sort_values(by=['振分先コード', '案件管理シート送信実施日'])


    # 結果の出力
    base_name = f'{YYYYMMDD}_案件管理シート送信実績取込用'
    save_filepath = serial_filepath(result_folder, base_name, '.csv')
    output_df.to_csv(save_filepath, index=False, encoding='CP932')
    
    print(f' >> {str(save_filepath)} を作成')


# ============================================
# メイン処理
# ============================================
if __name__ == '__main__':

    try:
        # # ==== フォルダ作成 ====
        # make_folders()

        # # ==== 案件管理シート作成 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 案件管理シート作成開始')
        # create_sheets()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 案件管理シート作成終了')

        # # ==== 案件管理シートチェック ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] チェック開始')
        # check_sheets_at_creation()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] チェック終了')

        # # ==== 送付対象クライアント一覧出力 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 送付対象クライアント一覧作成開始')
        # create_send_list()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 送付対象クライアント一覧作成終了')

        
        # # ==== 案件管理シートの送信実績取込用ファイル作成 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 案件管理シートの送信実績取込用ファイル作成開始')
        # create_send_jisseki()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 案件管理シートの送信実績取込用ファイル作成終了')

        pass # no excute



    except KeyboardInterrupt:
        print('\n⛔ 処理を中断しました')
    except Exception as e:
        import traceback
        err_name = type(e).__name__
        err_msg = str(e)
        err_detail = traceback.format_exc()
        print(f"\n⚠️ エラーが発生しました \nエラー名: {err_name}\n詳細: {err_msg}\n{err_detail}")