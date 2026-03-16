# -*- coding: utf-8 -*-
"""
Project: サンレモ成約捕捉
File: check_sheets.py
Description: 
    1. 作業フォルダの作成
    2. 案件管理シートのチェック・振分け
    3. 受領連絡宛先リストの作成
    4. 回収実績反映リストの作成
    5. 反響情報取込用ファイルの作成

Copyright (c) 2026 SCSK ServiceWare Corporation.
All rights reserved.

"""

import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import numpy as np
import datetime
import os
import re
import io
from msoffcrypto import OfficeFile
import shutil
from pathlib import Path
from itertools import count

from utils import (
    # 設定値
    set_config,
    # 独自エラー
    AppError, ZeroDataError, NotSelectError, ReferencePathError, NoSheetError,
    # 共通関数
    get_pw, select_folder, select_file, search_file, selectfile_to_df, serial_filepath, sanitize_filename, get_base_dir,
    create_name, create_filename, exist_folders
)

# ============================================
# 各種変数・関数
# ============================================

# 日付文字列生成
YYYYMMDD = datetime.date.today().strftime("%Y%m%d")
YYYYMM = datetime.date.today().strftime("%Y%m")

# ==== 外部参照 ====
config = set_config()


# ============================================
# フォルダーの作成
# ============================================
ROOT_FOLDER = Path(config['ルートフォルダ'])
BASE_FOLDER = ROOT_FOLDER / '04_案件管理シート受理・取込み' 
STORE_FOLDER = BASE_FOLDER / '■チェック前案件管理シート格納'
WORKING_FOLDER = BASE_FOLDER / YYYYMMDD
CSV_FOLDER = WORKING_FOLDER / '01_CSV'
PROCESSED_FOLDER = WORKING_FOLDER / '02_チェック済み案件管理シート'
RESULT_FOLDER = WORKING_FOLDER / '99_処理結果'

# フォルダの作成（既存のものは触れない）
def make_folders():
    STORE_FOLDER.mkdir(parents=True, exist_ok=True)
    WORKING_FOLDER.mkdir(parents=True, exist_ok=True)
    CSV_FOLDER.mkdir(exist_ok=True)
    PROCESSED_FOLDER.mkdir(exist_ok=True)
    RESULT_FOLDER.mkdir(exist_ok=True)


# ============================================
# 列名の改行など置換　\n \\n \u200b
# ============================================
def clean_column_name(name):
    if not isinstance(name, str):
        return name
    # \\n (文字列としての改行), \n (実際の改行), \u200b (ゼロ幅スペース) を除去
    return name.replace('\\n', '').replace('\n', '').replace('\u200b', '').strip()
    
# ============================================
# 案件管理シートチェッカー
# ============================================
class SheetChecker:
    def __init__(self, file_path):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        
        self.all_cols = config['案件管理シート']['全体範囲']
        self.input_cols = config['案件管理シート']['入力範囲']
        self.void_cols = config['案件管理シート']['空列範囲']
        self.sheet_name = config['案件管理シート']['シート名']

        # 後続のチェックで共通利用するデータ
        self.furiwakesaki_code = ''
        self.kigyo_name = ''
        self.furiwakesaki_name = ''
        self.target_date = ''

        self.pw = None
        self.wb = None
        self.ws = None
        self.df = None

        self.last_row = 9
        self.row_count = 0

    def get_info(self):
        """企業情報の取得"""
        # 振分先コード取得
        digit_matches = re.findall(r"\d{11}", self.file_name) # 11桁の数字
        if not digit_matches: return 'CodeErr' # 振分先コードが取得できない
        self.furiwakesaki_code = digit_matches[0]
        
        # PW取得
        self.pw = get_pw(self.furiwakesaki_code)
        if not self.pw: return 'PwErr' # PWが存在しない

        # pw開封してdf, wb/wsをそれぞれ取得
        try:
            with open(self.file_path, "rb") as f:
                office_file = OfficeFile(f)
                decrypted_data = io.BytesIO()
                office_file.load_key(password=self.pw)
                office_file.decrypt(decrypted_data)
                buf = decrypted_data.getvalue()

                try:
                    # dfの取得
                    # self.df = pd.read_excel(io.BytesIO(buf), dtype=str, usecols=self.all_cols, header=6, skiprows=[7]).fillna('')
                    self.df = pd.read_excel(io.BytesIO(buf), dtype=str, header=6, skiprows=[7]).fillna('')
                    self.df = self.df.iloc[:, 1:] # 2列目以降指定
                    self.df = self.df.map(lambda x: x.strip() if isinstance(x, str) else x) # データクレンジング

                except Exception as e:
                    print(e)
                    return 'DataFrameErr' # DFオブジェクトが作成できない
                
                try:
                    self.wb = load_workbook(io.BytesIO(buf), data_only=True)
                    self.ws = self.wb[self.sheet_name]
                    
                except Exception as e:
                    print(e)
                    return 'OpenpyxlErr' # WBオブジェクトが作成できない
                
        except Exception as e:
            print(e)
            return 'OpenErr' # エクセルが開けない or DF.WBオブジェクトが作成できない
        
        
        # 企業名、振分先名、対象月の取得
        self.target_date = self.ws['B1'].value
        if not self.target_date: return 'DateErr' # 対象月が存在しない

        pattern = r"^([^（]+)(?:（(.+)分）)?$"
        match = re.search(pattern, self.ws['D3'].value or "")
        if not match: return 'NameErr' # 企業名が取得できない
        self.kigyo_name = match.group(1).rstrip("様")
        self.furiwakesaki_name = match.group(2) or '' # なければNone

        # 件数、最終行
        self.row_count = len(self.df)
        self.last_row = 8 + len(self.df)


    # --- ファイル名 ---
    def check_filename(self): 
        """ファイル名が正しい"""
        try:
            correct = create_filename(
                self.furiwakesaki_code,
                self.kigyo_name,
                self.furiwakesaki_name
            )
            return self.file_name == correct
        except Exception as e:
            return False 


    def check_extension(self): 
        """拡張子がxlsx"""
        try:
            return self.file_name.endswith('.xlsx')            
        except Exception as e:
            return False 
            
    def check_month(self):
        """B1の日付が今月"""
        try:
            now = datetime.datetime.now()
            return self.target_date.year == now.year and self.target_date.month == now.month
        except Exception as e:
            return False 
            

    def check_columns_del(self):
        """列が削除されていない"""
        try:
            actual_cols = {clean_column_name(col) for col in self.df.columns}
            expected_cols = {clean_column_name(col) for col in config['案件管理シート']['列名']}
            return expected_cols.issubset(actual_cols)  # 期待列がすべて存在すればOK
        except Exception as e:
            return False 
            

    def check_columns_add(self):
        """列が追加されていない"""
        try:
            actual_cols = {clean_column_name(col) for col in self.df.columns}
            expected_cols = {clean_column_name(col) for col in config['案件管理シート']['列名']}
            return actual_cols.issubset(expected_cols)  # 実際の列がすべて期待列内ならOK
        except Exception as e:
            return False 
            

    def check_empty(self, col_name='依頼番号'):
        """最終行まで空欄がない"""
        try:
            return not (self.df[col_name] == '').any()    
        except Exception as e:
            return False 
                
    def check_status(self, col_name='②ご商談状況​'):
        """想定外の商談状況がない"""
        try:
            status_list = config['案件管理シート']['商談状況']
            return bool(self.df.loc[self.df[col_name] != '', col_name].isin(status_list).all()) # np.true_ 対策
        except Exception as e:
            return False 
            

    def check_is_date(self, col_name='日付'):
        """日付列が有効な日付型"""
        try:
            converted = pd.to_datetime(self.df[col_name], errors='coerce')
            return (converted.notna() | (self.df[col_name] == '')).all()    
        except Exception as e:
            return False 
            
    def check_is_num(self, col_name='重複\n※更新先の依頼番号を入力'):
        """値が数値文字列"""
        try:
            target = self.df.loc[self.df[col_name] != '', col_name]
            return bool(target.str.isdigit().all()) # np.true_ 対策
        except Exception as e:
            return False 
            

    def detect_garbled_text(self, text):
        """[関数] 文字化け検知"""
        if text == "": return False

        # 1. 不明な文字への置換（致命的な欠落）
        # システムが解釈不能で '?' や '' (U+FFFD) になったもの
        if '?' in text or '\ufffd' in text:
            return True

        # 2. UTF-8をLatin-1(ISO-8859-1)等で誤認した際の記号（Ã, Å, æ 等）
        # 日本語の文章にこれらのラテン文字が混じることはまずありません
        latin_garble_pattern = re.compile(r'[ÃÂÅÐÑÒÓÔÕÖ×àáâãäåæçèéêëìíîïðñòóôõö÷øùúûüýþÿ]')
        if latin_garble_pattern.search(text):
            return True

        # 3. 制御文字の混入
        # 文字化けの結果、本来Excelに入らないはずの制御文字(改行・タブ以外)が混じることがあります
        # C0制御文字（0x00-0x1F）のうち、タブ(\t), 改行(\n, \r) 以外が含まれるか
        control_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
        if control_chars.search(text):
            return True

        # 4. エンコードの整合性チェック（最重要）
        # 日本のCL環境(Windows/Excel)で作成されたなら、CP932(Shift-JIS拡張)でエンコードできるはず
        # これに失敗する場合、UTF-8の特殊文字が化けているか、保存形式が壊れています
        try:
            text.encode('cp932')
        except UnicodeEncodeError:
            # CP932で扱えない文字（環境依存文字の一部や、化けた文字）が含まれる
            return True

        # 5. 「縺」などの特定の化け頻出漢字（Shift-JISをUTF-8で開いたパターン）
        # 意味をなさない「縺」「縺」「縺」などの連続を検知
        sjis_as_utf8_pattern = re.compile(r'[縺縅縈縋縐縑縒縓縔縕縖縗縘縚縜]')
        if sjis_as_utf8_pattern.search(text):
            return True

        return False

    def check_mojibake(self):
        """シート内容全体に文字化けがない"""
        try:
            return not self.df.map(self.detect_garbled_text).any().any()
        except Exception as e:
            return False 
            
    def check_cell_protection(self):
        """CL入力欄以外ロックされている"""
        try:
            if not self.ws.protection.sheet: return False

            # configから数値化
            all_start_idx = column_index_from_string(config['案件管理シート']['全体範囲'].split(':')[0])
            all_end_idx = column_index_from_string(config['案件管理シート']['全体範囲'].split(':')[1])
            input_start_idx = column_index_from_string(config['案件管理シート']['入力範囲'].split(':')[0])
            input_end_idx = column_index_from_string(config['案件管理シート']['入力範囲'].split(':')[1])

            # 行ごとに確認
            for row in self.ws.iter_rows(min_row=9, max_row=self.last_row, min_col=all_start_idx, max_col=all_end_idx):
                for cell in row:
                    col_idx = cell.column
                    is_locked = cell.protection.locked
                    if  input_start_idx <= col_idx <= input_end_idx:
                        if is_locked: return False
                    else:
                        if not is_locked: return False
            
            return True
        except Exception as e:
            return False 
        

    def check_input_range(self):
        """A9以下、AA9以下、最終行以降に値が入力されていない"""
        try:
            all_start_idx = column_index_from_string(config['案件管理シート']['全体範囲'].split(':')[0])
            all_end_idx = column_index_from_string(config['案件管理シート']['全体範囲'].split(':')[1])

            left_cell = self.ws.cell(9, all_start_idx - 1).value
            right_cell = self.ws.cell(9, all_end_idx + 1).value
            under_cells = [self.ws.cell(self.last_row + 1, col).value for col in range(all_start_idx, all_end_idx + 1)]

            return all([
                left_cell is None or left_cell == '',
                right_cell is None or right_cell == '',
                all([val is None or val == '' for val in under_cells])
            ])
        except Exception as e:
            return False 
        
# ============================================
# 結果に応じてファイルを格納
# ============================================
def move_file(src_path, result:bool):
    if result:
        dst_path = PROCESSED_FOLDER / '不備なし' / os.path.basename(src_path)
    else:
        dst_path = PROCESSED_FOLDER / '不備あり' / os.path.basename(src_path)
    # フォルダ作成 
    dst_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(src_path), str(dst_path))


# ============================================
# 格納フォルダ内の案件管理シートをすべてチェックし、結果出力
# ============================================
def check_sheets_on_receipt():

    exist_folders([BASE_FOLDER, STORE_FOLDER, WORKING_FOLDER, CSV_FOLDER, PROCESSED_FOLDER, RESULT_FOLDER])

    result = []

    # 格納フォルダから案件管理シートのパスを取得
    files = [f for f in STORE_FOLDER.glob("*.xlsx") if not f.name.startswith('~$')]
    if len(files) == 0: 
        raise ZeroDataError(f'案件管理シートが格納されていません')
    
    
    # 送付先リストを取得 -> R営業担当者氏名、Rクラサポ担当者氏名
    send_df = selectfile_to_df("案件管理シート送付先リスト.csv", CSV_FOLDER)

    ok_count = 0
    ng_count = 0
    ok_file_count = 0
    ng_file_count = 0
    
    # 格納フォルダ内の案件管理シートを取得
    for i, file in enumerate(files):
        # 進捗
        print(f'[{i+1}/{len(files)}] 処理開始: {str(file.name)}')

        # チェック
        checker = SheetChecker(str(file))
        res = checker.get_info()

        report = {
            '振分先コード': checker.furiwakesaki_code or "",
            '企業名': checker.kigyo_name or "",
            '振分先名': checker.furiwakesaki_name or "",
            'R営業担当者氏名': '',
            'Rクラサポ担当者氏名': '',
            'トータルの不備チェック結果': '',
            '不備内容(不備あり時)1': '',
            '不備内容(不備あり時)2': '',
            'ファイル名': str(file),
            'ファイル形式チェック': '',
            '過去分ファイルチェック': '',
            'CL入力欄削除列チェック': '',
            'CL入力欄追加列チェック': '',
            '依頼番号空白チェック': '',
            '商談ステータスチェック': '',
            '日付データ型チェック': '',
            '重複依頼番号データ型チェック': '',
            'CL入力欄文字化けチェック': ''
        }

        # 担当者名取得
        send_row = send_df[send_df['振分先コード'] == checker.furiwakesaki_code]
        if not send_row.empty:
            report['R営業担当者氏名'] = send_row['R営業担当者氏名'].values[0]
            report['Rクラサポ担当者氏名'] = send_row['Rクラサポ担当者氏名'].values[0]


        match res:
            case 'CodeErr':
                report['トータルの不備チェック結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ 振分先コードを取得できません')
                ng_count += 1
                continue
            case 'PwErr':
                report['トータルの不備チェック結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ パスワードを取得できません')
                ng_count += 1
                continue
            case 'OpenErr':
                report['トータルの不備チェック結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ ファイルを開けません')
                ng_count += 1
                continue
            case 'DataFrameErr':
                report['トータルの不備チェック結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ シート内の情報を取得できません')
                ng_count += 1
                continue
            case 'OpenpyxlErr':
                report['トータルの不備チェック結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ シート内の情報を取得できません')
                ng_count += 1
                continue
            case 'NameErr':
                report['トータルの不備チェック結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ 社名を取得できません')
                ng_count += 1
                continue
            case None:
                pass
  
        
        report['ファイル形式チェック'] = checker.check_extension()
        report['過去分ファイルチェック'] = checker.check_month()
        report['CL入力欄削除列チェック'] = checker.check_columns_del()
        report['CL入力欄追加列チェック'] = checker.check_columns_add()
        report['依頼番号空白チェック'] = checker.check_empty()
        report['商談ステータスチェック'] = checker.check_status()
        report['日付データ型チェック'] = all([
            checker.check_is_date('①最終更新日​'),
            checker.check_is_date('③初回来場日​'),
            checker.check_is_date('④契約予定日​'),
        ])
        report['重複依頼番号データ型チェック'] = checker.check_is_num()
        report['CL入力欄文字化けチェック'] = checker.check_mojibake()

        # トータルで
        check_targets = [        
            report['ファイル形式チェック'],
            report['過去分ファイルチェック'],
            report['CL入力欄削除列チェック'],
            report['CL入力欄追加列チェック'] ,
            report['依頼番号空白チェック'],
            report['商談ステータスチェック'], # npbool
            report['日付データ型チェック'],
            report['重複依頼番号データ型チェック'], # npbool
            report['CL入力欄文字化けチェック']
        ]


        # 全部TrueならTrue、一つでもFalseがあればFalse
        report['トータルの不備チェック結果'] = 'OK' if all(check_targets) else 'NG'

        # 不備内容出力
        error_messages = config['不備']

        # False（不備）があったメッセージだけを抽出
        error_texts = [msg for key, msg in error_messages.items() if not report.get(key)]

        # メッセージの数に応じて text, text2 に振り分け
        # 最初の5つまで
        report['不備内容(不備あり時)1'] = " / ".join(error_texts[:5]) if error_texts else ""

        # 6つ目以降がある場合
        if len(error_texts) >= 6:
            report['不備内容(不備あり時)2'] = " / ".join(error_texts[5:])
        else:
            report['不備内容(不備あり時)2'] = ""

        result.append(report)
        print(' ✔  完了')
        
        ok_count += 1


        # ファイル移動
        if report['トータルの不備チェック結果'] == 'OK':
            move_file(str(file), True)
            ok_file_count += 1
        else:
            move_file(str(file), False)
            ng_file_count += 1


    print(f' ✅ チェック完了: {ok_count} | チェック不可: {ng_count} | 合計: {len(files)}')

    # 結果を出力
    result_df = pd.DataFrame(result)
    result_df = result_df.replace(True, 'OK').replace(False, 'NG')

    # チェック結果テンプレート
    template_filepath = Path(config['ファイルパス']['受理不備チェック結果FMT'])

    # 保存ファイル名(すでにある場合はリネームして新規で作成)
    base_name = f"{YYYYMMDD}_案件管理シート受理不備チェック結果"
    save_filepath = serial_filepath(RESULT_FOLDER, base_name, '.xlsx')
    shutil.copy(template_filepath, save_filepath)

    # 全体の結果書き込み

    # pandasの結果を流し込む
    with xw.App(visible=False) as app:
        wb = xw.Book((save_filepath))
        ws = wb.sheets[0]  # 1番左のシートを指定
        ws.range("A:A").number_format = '@'
        ws.range('A9').options(index=False, header=False).value = result_df

        ws.range('C3').value = len(files)
        ws.range('C4').value = ok_file_count
        ws.range('C5').value = ng_file_count
        
        # 保存
        wb.save()
        wb.close()

    print(f' >> {str(save_filepath)} を作成')



# ============================================
# 案件管理シートの回収実績取込用ファイル作成
# ============================================
def create_receive_jisseki():

    exist_folders([BASE_FOLDER, CSV_FOLDER])

    # 結果取得
    title = f'案件管理シート受理不備チェック結果.xlsxを選択してください'
    print(title)
    result_filepath = select_file(title=title, file_types=[('EXCELファイル', '*.xlsx')], initial_dir=BASE_FOLDER)
    result_folder = result_filepath.parent

    result_df = pd.read_excel(result_filepath, engine='openpyxl', skiprows=7, usecols='A:R', dtype=str).fillna('')
    result_dict = result_df.set_index('振分先コード').to_dict(orient='index')

    # csvファイルの取得
    sojushin_df = selectfile_to_df("案件管理シート送受信実績一覧.csv", CSV_FOLDER).fillna('')
    group = sojushin_df.groupby('振分先コード')

    # 日付データの整形
    sojushin_df['案件管理シート送信実施日'] = pd.to_datetime(sojushin_df['案件管理シート送信実施日']).dt.strftime('%Y/%m/%d')
    sojushin_df['案件管理シート回収日'] = pd.to_datetime(sojushin_df['案件管理シート回収日']).dt.strftime('%Y/%m/%d')
    
    # 今日の日付
    today_str = datetime.date.today().strftime('%Y/%m/%d')
    updated_dfs = []

    for code, df in group:
        # Excelの結果から該当コードの情報を取得
        row = result_dict.get(str(code)) # 型不一致を防ぐためstr化
        
        if row is None:
            # 結果Excelに載っていないコードは除外
            continue
        
        # 辞書から値を取得（.iloc[0]は不要）
        is_ok = row.get('トータルの不備チェック結果') == 'OK'
        hubi1 = row.get('不備内容(不備あり時)1', '')
        hubi2 = row.get('不備内容(不備あり時)2', '')
        hubi_text = f"{hubi1}/{hubi2}" if hubi2 else hubi1

        temp_df = df.copy()
        
        # 最終行（最新の履歴）を更新
        if is_ok:
            temp_df.iloc[-1, 3] = today_str
            temp_df.iloc[-1, 4] = '不備なし'
            temp_df.iloc[-1, 5] = ''
        else:
            temp_df.iloc[-1, 3] = today_str
            temp_df.iloc[-1, 4] = '不備解消中'
            temp_df.iloc[-1, 5] = hubi_text

        updated_dfs.append(temp_df)

    # 5. 結果の出力
    if updated_dfs:
        output_df = pd.concat(updated_dfs, ignore_index=True)
        # 最後にコードと日付で並び替え
        output_df = output_df.sort_values(by=['振分先コード', '案件管理シート送信実施日'])
        
        base_name = f'{YYYYMMDD}_案件管理シート回収実績取込'
        save_filepath = serial_filepath(result_folder, base_name, '.csv')
        
        #
        output_df.to_csv(save_filepath, index=False, encoding='CP932')
        print(f' >> {save_filepath.name} を作成しました')

    else:
        print(f' 案件管理シート回収実績取込.csv を作成できませんでした')


# ============================================
# CLへの受領メール宛先リスト作成
# ============================================
def create_mail_list():

    # == チェック結果の取得 ==
    title = f'案件管理シート受理不備チェック結果.xlsxを選択してください'
    print(title)
    result_filepath = select_file(title=title, file_types=[('EXCELファイル', '*.xlsx')], initial_dir=BASE_FOLDER)
    result_folder = result_filepath.parent

    result_df = pd.read_excel(result_filepath, engine='openpyxl', skiprows=7, usecols='A:R', dtype=str).fillna('')

    send_df = selectfile_to_df("案件管理シート送付先リスト.csv", BASE_FOLDER)
    person_df = selectfile_to_df("案件管理担当者リスト.csv", BASE_FOLDER)

    # --- 1. 結合 --- 
    merged_df1 = pd.merge(
        send_df[['振分先コード', '案件管理主体の振分先コード', '企業名', '振分先名']], 
        person_df[['振分先コード', '担当者部署', '担当者氏名', '担当者メールアドレス']], 
        left_on='案件管理主体の振分先コード', 
        right_on='振分先コード', 
        how='inner'
    )
    # 列整理
    merged_df1 = merged_df1.rename(columns={'振分先コード_x': '振分先コード'})
    merged_df1 = merged_df1.drop(columns=['案件管理主体の振分先コード', '振分先コード_y'])

    # 出力
    base_name = f"{YYYYMMDD}_案件管理シート全送付先リスト"
    save_filepath = next(
        path for i in count()
        if not (path := result_folder / (f"{base_name}_{i}.csv" if i > 0 else f"{base_name}.csv")).exists()
    )
    merged_df1.to_csv(save_filepath, index=False, encoding="CP932")


    # OKの宛先リスト
    ok_df = result_df[result_df["トータルの不備チェック結果"] == "OK"]
    if len(ok_df) > 0:
        merged_df_ok = pd.merge(
            merged_df1, 
            ok_df[['振分先コード']], 
            left_on='振分先コード', 
            right_on='振分先コード', 
            how='inner'
        )

        #  --- 4. 整形 --- 
        def create_name_format(row):
            kigyo = row['企業名']
            furi = row['振分先名']
            busho = row['担当者部署']
            shimei = row['担当者氏名']
            
            # 振分先名がある場合だけカッコを付ける
            furi_part = f" ({furi}分)" if furi != "" else ""
            
            # 全体を組み立てる
            return f"{kigyo}{furi_part} {busho} {shimei}"

        # 新しいデータフレームを作成
        output_df = pd.DataFrame()
        output_df['名前'] = merged_df_ok.apply(create_name_format, axis=1)
        output_df['メールアドレス'] = merged_df_ok['担当者メールアドレス']
        output_df['グループ'] = f'{YYYYMMDD}_案件管理シート受理完了'
        output_df['コメント'] = ""
        output_df['属性1'] = ""
        output_df['属性2'] = ""

        # 並び替え
        cols = ['グループ', '名前', 'メールアドレス', 'コメント', '属性1', '属性2']
        output_df = output_df[cols]


        #  --- 5. 出力 --- 
        base_name = f"{YYYYMMDD}_案件管理シート受領完了メール宛先リスト"
        save_filepath = serial_filepath(result_folder, base_name, '.csv')
        output_df.to_csv(save_filepath, index=False, encoding="CP932")

        print(f' >> {str(save_filepath)} を作成')

    ng_df = result_df[result_df["トータルの不備チェック結果"] == "NG"]
    if len(ng_df) > 0:
        merged_df_ng = pd.merge(
            merged_df1, 
            ng_df[['振分先コード', '不備内容(不備あり時)1', '不備内容(不備あり時)2']], 
            left_on='振分先コード', 
            right_on='振分先コード', 
            how='inner'
        )
        #  --- 4. 整形 --- 
        def create_name_format(row):
            kigyo = row['企業名']
            furi = row['振分先名']
            busho = row['担当者部署']
            shimei = row['担当者氏名']
            
            # 振分先名がある場合だけカッコを付ける
            furi_part = f" ({furi}分)" if furi != "" else ""
            
            # 全体を組み立てる
            return f"{kigyo}{furi_part} {busho} {shimei}"
        
        # 新しいデータフレームを作成
        output_df = pd.DataFrame()
        output_df['名前'] = merged_df_ng.apply(create_name_format, axis=1)
        output_df['メールアドレス'] = merged_df_ng['担当者メールアドレス']
        output_df['グループ'] = f'{YYYYMMDD}_案件管理シート不備差戻'
        output_df['コメント'] = ""
        output_df['属性1'] = merged_df_ng['不備内容(不備あり時)1']
        output_df['属性2'] = merged_df_ng['不備内容(不備あり時)2']
        # 並び替え
        cols = ['グループ', '名前', 'メールアドレス', 'コメント', '属性1', '属性2']
        output_df = output_df[cols]


        #  --- 5. 出力 --- 
        base_name = f"{YYYYMMDD}_案件管理シート不備差戻メール宛先リスト"
        save_filepath = serial_filepath(result_folder, base_name, '.csv')
        output_df.to_csv(save_filepath, index=False, encoding="CP932")


        print(f' >> {str(save_filepath)} を作成')


# ============================================
# 反響内容取込用ファイル作成
# ============================================
def create_input_data_hankyo():

    # チェック済みの案件管理シート格納フォルダ指定
    title = f'対象の日付の作業フォルダを選択してください(YYYYMMDD)'
    print(title)
    working_filepath = select_folder(title=title, initial_dir=BASE_FOLDER)

    processed_folder = working_filepath / '02_チェック済み案件管理シート'
    result_folder = working_filepath / '99_処理結果'
    complete_folder = processed_folder / '不備なし'

    files =  [f for f in complete_folder.glob("*.xlsx") if not f.name.startswith('~$')]
    if not files: raise FileNotFoundError('不備なしの案件管理シートがありません')
    
    concat_dfs = []
    err_count = 0
    ok_count = 0

    for k, file in enumerate(files):

        print(f'[{k+1}/{len(files)}] 処理開始: {file}')

        try: 
            # ファイル名から振分先コード取得
            match = re.search(r'\d{11}', file.name)
            code = match.group() if match else None
            pw = get_pw(code)
            if not pw:
                print(' ⚠️ pwなし')
                err_count += 1
                continue

            with xw.App(visible=False) as app:
                wb = app.books.open(file, password=pw)
                ws = wb.sheets[config['案件管理シート']['シート名']]

                last_row = ws.used_range.last_cell.row
                left_col, right_col = config['案件管理シート']['全体範囲'].split(':')

                data = ws.range(f'{left_col}9:{right_col}{last_row}').value

                wb.close()
            
            df = pd.DataFrame(data, columns=None)
            
            # 整形(空行削除)
            df = df.replace('', np.nan).dropna(how='all', axis=0).fillna('')
            
            # F列(5) と P～V列(15～21) を取得
            selected_df = df.iloc[:, [4, 14, 15, 16, 17, 18, 19]]
            concat_dfs.append(selected_df)

        except Exception as e:
            print(' ⚠️ データ取得時エラー')
            err_count += 1
            continue
        
        print(' ✔  完了')
        ok_count += 1

    print(f' ✅ 取得完了: {ok_count} | 作成エラー: {err_count} | 合計: {len(files)}')


    # まとめる
    if not concat_dfs:
        raise ValueError("有効データがありません")
    
    output_df = pd.concat(concat_dfs, ignore_index=True)
    output_df.columns = [
        '依頼番号',
        '最終更新日（クライアント）',
        'ご商談状況',
        '初回来場日',
        '契約予定日',
        '営業担当者氏名　※任意',
        '重複　※更新先の依頼番号を入力'
    ]

    output_df['最終更新日（クライアント）'] = pd.to_datetime(output_df['最終更新日（クライアント）']).dt.date
    output_df['初回来場日'] = pd.to_datetime(output_df['初回来場日']).dt.date
    output_df['契約予定日'] = pd.to_datetime(output_df['契約予定日']).dt.date

    # 出力
    base_name = f"{YYYYMMDD}_案件管理シート反響管理アプリ反映"
    save_filepath = serial_filepath(result_folder, base_name, '.csv')
    output_df.to_csv(save_filepath, index=False, encoding="CP932")
    
    print(f" >> {save_filepath} を作成")


# --- 実行例 ---
if __name__ == "__main__":
        
    try:
        # ==== フォルダ作成 ====
        # make_folders()

        # # ==== 案件管理シートチェック ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] チェック開始')
        # check_sheets_on_receipt()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] チェック終了')


        # # ==== 送受信実績ファイル作成 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 取込用ファイル作成開始')
        # create_receive_jisseki()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 取込用ファイル作成終了')


        # # ==== 宛先送付リストファイル作成 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 送付リスト作成開始')
        # create_mail_list()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 送付リスト作成終了')


        # # ==== 反響取込用ファイル作成 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 反映用ファイル作成開始')
        # create_input_data_hankyo()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 反映用ファイル作成終了')

        pass

    except KeyboardInterrupt:
        print('\n⛔ 処理を中断しました')    
    except Exception as e:
        import traceback
        err_name = type(e).__name__
        err_msg = str(e)
        err_detail = traceback.format_exc()
        print(f"\n⚠️ エラーが発生しました error.log を確認してください\nエラー名: {err_name}\n詳細: {err_msg}\n{err_detail}")