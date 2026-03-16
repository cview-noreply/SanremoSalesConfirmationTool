# -*- coding: utf-8 -*-
"""
Project: サンレモ成約捕捉
File: creat_alert.py
Description: 
    1. 案件アラートの作成
    2. 作成した案件アラートのチェック(check_sheets.py のチェックルールを援用)、チェック結果の作成
    3. 送信対象クライアント一覧の作成

Copyright (c) 2026 SCSK ServiceWare Corporation.
All rights reserved.
"""

import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import datetime
import traceback
import os
import re
import io
from msoffcrypto import OfficeFile
import shutil
from pathlib import Path

from utils import (
    # 設定値
    set_config,
    # 独自エラー
    AppError, ZeroDataError, NotSelectError, ReferencePathError, NoSheetError,
    # 共通関数
    get_pw, select_folder, select_file, search_file, selectfile_to_df, serial_filepath, sanitize_filename, get_base_dir,
    create_name, create_filename, create_filename_alert, exist_folders
)


# ==== 日付文字列生成 ====
YYYYMMDD = datetime.date.today().strftime("%Y%m%d")
YYYYMM = datetime.date.today().strftime("%Y%m")

# ==== 外部参照 ====
config = set_config()


# ============================================
# 案件アラートチェッカー
# ============================================
class AlertChecker:
    def __init__(self, file_path):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        
        self.sheet_name = config['案件アラート']['シート名']
        self.all_cols = config['案件アラート']['全体範囲']

        # 後続のチェックで共通利用するデータ
        self.furiwakesaki_code = ''
        self.kigyo_name = ''
        self.furiwakesaki_name = ''

        self.pw = None
        self.wb = None
        self.ws = None
        self.df = None

        self.last_row = 9
        self.row_count = 0

    def get_info(self):
        """企業情報の取得"""
        # 振分先コード取得
        digit_matches = re.findall(r"\d{11}", self.file_name) # 11桁の数字

        if digit_matches:
            self.furiwakesaki_code = digit_matches[0]
            
            # PW取得
            self.pw = get_pw(self.furiwakesaki_code)
            if not self.pw: return 'PwErr' # PWが存在しない

        else:
            # ファイル名から企業名と振分先を取得
            pattern = r"案件アラート_(.+?)様(?:（(.+?)分）)?"
            match = re.search(pattern, self.file_name)
            if not match: return 'NameErr'

            self.kigyo_name = match.group(1)
            self.furiwakesaki_name = match.group(2) or ''


            # PW一覧内を企業名と振分先名で探索
            pw_filepath = config['ファイルパス']['PW一覧']
            pw_df = pd.read_excel(pw_filepath, dtype=str).fillna('')

            pw_info = pw_df.loc[
                (pw_df['企業名'] == self.kigyo_name) &
                (pw_df['振分先名'] == self.furiwakesaki_name),
                ['PW', '振分先コード']
            ]

            if pw_info.empty or len(pw_info) > 1:
                return 'PwErr'
            
            self.pw = pw_info.iloc[0]['PW']
            self.furiwakesaki_code = pw_info.iloc[0]['振分先コード']


        # pw開封してdf, wb/wsをそれぞれ取得
        try:
            with open(self.file_path, "rb") as f:
                office_file = OfficeFile(f)
                decrypted_data = io.BytesIO()
                office_file.load_key(password=self.pw)
                office_file.decrypt(decrypted_data)
                buf = decrypted_data.getvalue()

                # dfの取得
                self.df = pd.read_excel(io.BytesIO(buf), dtype=str, sheet_name=self.sheet_name, usecols=self.all_cols, skiprows=4).fillna('')
                self.df = self.df.map(lambda x: x.strip() if isinstance(x, str) else x) # データクレンジング

                self.wb = load_workbook(io.BytesIO(buf), data_only=True)
                self.ws = self.wb[self.sheet_name]

        except Exception as e:
            return 'OpenErr' # エクセルが開けない or DF.WBオブジェクトが作成できない
        
        # 企業名、振分先名、対象月の取得
        pattern = r"^([^（]+)(?:（(.+)分）)?$"
        match = re.search(pattern, self.ws['D3'].value or "")
        if not match: return 'NameErr' # 企業名が取得できない
        self.kigyo_name = match.group(1).rstrip("様")
        self.furiwakesaki_name = match.group(2) or '' # なければNone

        # 件数、最終行
        self.row_count = len(self.df)
        self.last_row = 5 + len(self.df)


    # --- ファイル名 ---
    def check_filename(self): 
        """ファイル名が正しい"""
        correct = create_filename_alert(
            self.furiwakesaki_code,
            self.kigyo_name,
            self.furiwakesaki_name
        )
        return self.file_name == correct

    
    def check_input_range(self):
        """B6以下、Q6以下、最終行以降に値が入力されていない"""
        all_start_idx = column_index_from_string(config['案件アラート']['全体範囲'].split(':')[0])
        all_end_idx = column_index_from_string(config['案件アラート']['全体範囲'].split(':')[1])

        left_cell = self.ws.cell(6, all_start_idx - 1).value
        right_cell = self.ws.cell(6, all_end_idx + 1).value
        under_cells = [self.ws.cell(self.last_row + 1, col).value for col in range(all_start_idx, all_end_idx + 1)]

        return all([
            left_cell is None or left_cell == '',
            right_cell is None or right_cell == '',
            all([val is None or val == '' for val in under_cells])
        ])


# ============================================
# フォルダーの作成
# ============================================

ROOT_FOLDER = Path(config['ルートフォルダ'])
BASE_FOLDER = ROOT_FOLDER / '23_案件アラート表作成送付'
WORKING_FOLDER = BASE_FOLDER / YYYYMMDD
CSV_FOLDER = WORKING_FOLDER / '01_CSV'
SHEETS_FOLDER = WORKING_FOLDER / '02_作成済み案件アラート'
RESULT_FOLDER = WORKING_FOLDER / '99_処理結果'

def make_folders():
    WORKING_FOLDER.mkdir(parents=True, exist_ok=True)
    CSV_FOLDER.mkdir(exist_ok=True)
    SHEETS_FOLDER.mkdir(exist_ok=True)
    RESULT_FOLDER.mkdir(exist_ok=True)


# ============================================
# CL向け案件アラートの作成
# ============================================
def create_sheets():

    exist_folders([BASE_FOLDER, WORKING_FOLDER, CSV_FOLDER, SHEETS_FOLDER, RESULT_FOLDER])

    # 結果出力用
    result = []

    # テンプレートファイルを指定
    template_filepath = config['ファイルパス']['案件アラートFMT']
    target_sheet_name = config['案件アラート']['シート名']
    restart_interval = int(config['xw_RESTART_INTERVAL'])

    # CL向け案件アラート表一覧の読み込み
    df = selectfile_to_df('CL向け案件アラート表一覧.csv', CSV_FOLDER)
    if len(df) == 0:
        raise ZeroDataError('データが0件です')


    # データを振分先コードでグループ化する
    group = df.groupby('振分先コード')
    group_list = list(group)

    app = None
    wb = None
    ws = None

    def _boot_app_and_open_template():
        """Excelアプリを起動してテンプレートを開き (app, wb, ws) を返す"""
        _app = xw.App(visible=False)
        _app.screen_updating = False
        _app.display_alerts = False
        try:
            _wb = _app.books.open(template_filepath)
            _ws = _wb.sheets[target_sheet_name]
        except Exception:
            _app.quit()
            raise RuntimeError(f'テンプレートまたはシート({target_sheet_name})が見つかりません')
        return _app, _wb, _ws

    err_count = 0
    ok_count = 0
    exist_count = 0

    try:
        for i, (code, df) in enumerate(group_list):

            # ── Excelアプリの起動 or 定期再起動 ──────────────────────────
            # ループの先頭で判断し、wb/ws も同時に取得
            if app is None or (i > 0 and i % restart_interval == 0):
                if app is not None:
                    try:
                        wb.close()   # ★ app.quit() 前に必ず wb を閉じる
                    except Exception:
                        pass
                    app.quit()

                app, wb, ws = _boot_app_and_open_template()


            # ── 各種情報取得 ──────────────────────────────────────────────
            kigyo_nm = df['企業名'].iloc[0]
            furiwakesaki_nm = df['振分先名'].iloc[0]
            rows = len(df) # データ件数

            # 進捗表示
            print(f'[{i+1}/{len(group_list)}] 処理開始: {code}|{kigyo_nm}|{furiwakesaki_nm}')
            
            entry = {
                '振分先コード': code,
                '企業名': kigyo_nm,
                '振分先名': furiwakesaki_nm,
                'pw': '',
                'データ件数': rows,
                '作成日時': f'{datetime.datetime.now():%Y/%m/%d %H:%M:%S}',
                'ファイル名': '',
                'パス内禁止文字': ''
            }

            # ── PW取得 ───────────────────────────────────────────────────
            pw = get_pw(code)
            if not pw:
                entry['pw'] = '一覧内になし'
                result.append(entry)
                print(' ⚠️ pwなし')
                err_count += 1
                continue
            entry['pw'] = pw

            # ── 保存ファイル名作成 & 禁止文字チェック ────────────────────
            save_filename = create_filename_alert(code, kigyo_nm, furiwakesaki_nm)
            rep_filename = sanitize_filename(save_filename)
            if rep_filename != save_filename:
                save_filename = rep_filename
                entry['パス内禁止文字'] = 'あり'
            entry['ファイル名'] = save_filename

            # ── 作成済みスキップ ─────────────────────────────────────────
            save_filepath = SHEETS_FOLDER / save_filename
            if save_filepath.exists():
                entry['作成日時'] = '過去に作成済み'
                result.append(entry)
                print(' ⚠️ 作成済み')
                exist_count += 1
                continue


            # ── Excel書き込み ────────────────────────────────────────────
            ws.range('D2').value = code
            ws.range('D3').value = create_name(kigyo_nm, furiwakesaki_nm)
            
            # データの並び替え
            df = df[[
                'ご確認ご依頼事項', '概要', '資料請求日', '反響種別', 'モデルハウス名', '依頼番号', '名前（漢字）', '名前（ふりがな）', '年齢',
                '郵便番号', '住所（都道府県）', '住所（市区町村以降）', 'E-mailアドレス', '電話番号', '建築予定地（名称）', 'お客様来場日（来場キャンペーン応募情報）', 'お客様来場時期（ヒアリング経由）', 'お客様申告契約日', 
                '成約時不備チェック', '着工時不備チェック']].copy()
            df['成約時不備チェック'] = df['成約時不備チェック'].replace({
                '未チェック': '',
                '不備確認中': '不備確認依頼中',
                '不備なし': '提出・チェック完了'
            })
            df['着工時不備チェック'] = df['着工時不備チェック'].replace({
                '未チェック': '',
                '不備確認中': '不備確認依頼中',
                '不備なし': '提出・チェック完了',
                '個別請求先確認中': '提出・チェック完了'
            })


            # データ貼り付け
            ws.range("E:E").number_format = '@'
            ws.range('C6').options(index=False, header=False).value = df


            # SaveAs でファイルを書き出した後、wb は「別名で開いている状態」になる。
            # 次ループでテンプレートとして再利用するため SaveAs 後に閉じ、
            # テンプレートを再度開いて wb/ws を更新する。
            wb.api.SaveAs(str(save_filepath), Password=pw)
            wb.close()
            wb = app.books.open(template_filepath)
            ws = wb.sheets[target_sheet_name]

            # ── 作成結果追加 ────────────────────────────────────────────
            result.append(entry)
            print(' ✔  完了')
            ok_count += 1

    finally:
        if app is not None:
            try:
                wb.close()
            except Exception:
                pass
            app.quit()

    print(f' ✅ 作成完了: {ok_count} | 作成済み: {exist_count} | 作成エラー: {err_count} | 合計: {len(group_list)}')

    # 結果の出力
    result_df = pd.DataFrame(result)
    result_df.index = range(1, len(result_df) + 1) # インデックスを1始まりに
    base_name = f'{YYYYMMDD}_案件アラート作成結果メモ'
    save_filepath = serial_filepath(RESULT_FOLDER, base_name, '.xlsx')
    result_df.to_excel(save_filepath)

    print(f' >> {save_filepath} を作成')

# ============================================
# CL向け案件アラートのチェック
# ============================================
def check_sheets_at_creation():
    
    exist_folders([BASE_FOLDER])

    dialog_title = f'対象の日付の作業フォルダを選択してください(YYYYMMDD)'
    print(dialog_title)
    selected_working_folder = select_folder(title=dialog_title, initial_dir=BASE_FOLDER)

    data_df = selectfile_to_df('CL向け案件アラート表一覧.csv', selected_working_folder)
    if len(data_df) == 0: raise ZeroDataError('データが0件です')
        
    count_dict = data_df['振分先コード'].value_counts().to_dict()
    count_dict_len = len(count_dict)

    
    target_folder = selected_working_folder / '02_作成済み案件アラート'
    result_folder = selected_working_folder / '99_処理結果'

    target_files = [f for f in target_folder.glob("*.xlsx") if not f.name.startswith('~$')]
    target_files_len = len(target_files)

    if target_files_len == 0:
        raise ZeroDataError(f'02_作成済み案件アラートが空です')

    result = []
    ng_count = 0
    ok_count = 0

    for j, file in enumerate(target_files):
        # 進捗
        print(f'[{j+1}/{len(target_files)}] 処理開始: {str(file.name)}')

        # チェック用クラスの作成
        checker = AlertChecker(str(file))
        # 情報の取得
        res = checker.get_info()

        report = {
            '振分先コード': checker.furiwakesaki_code or "",
            '企業名': checker.kigyo_name or "",
            '振分先名': checker.furiwakesaki_name or "",
            'ファイル名': str(file.name),
            'トータルの不備結果': '',
            'PWチェック': '',
            'データ貼り付け位置チェック': '',
            'ファイル内反響数チェック': '',
            'ファイル内反響数': '',
            '一覧データ内の反響数': '',
            'ファイル名称チェック': ''   
        }
        
        match res:
            case 'CodeErr':
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ 振分先コードを取得できません')
                ng_count += 1
                continue
            case 'PwErr':
                report['PWチェック'] = 'NG'
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ パスワードを取得できません')
                ng_count += 1
                continue
            case 'OpenErr':
                report['PWチェック'] = 'NG'
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ ファイルを開けません')
                ng_count += 1
                continue
            case 'NameErr':
                report['トータルの不備結果'] = 'NG'
                result.append(report)
                print(f' ⚠︎ 社名を取得できません')
                ng_count += 1
                continue
            case None:        
                report['PWチェック'] = 'OK'
    
        report['データ貼り付け位置チェック'] = checker.check_input_range()
        report['ファイル内反響数'] = checker.row_count
        report['一覧データ内の反響数'] = count_dict.get(checker.furiwakesaki_code, 0)
        report['ファイル内反響数チェック'] = checker.row_count == count_dict.get(checker.furiwakesaki_code, 0)
        report['ファイル名称チェック'] = checker.check_filename()
        # トータルで
        check_targets = [
            report['データ貼り付け位置チェック'],
            report['ファイル内反響数チェック'],
            report['ファイル名称チェック']
        ]

        # 全部TrueならTrue、一つでもFalseがあればFalse
        report['トータルの不備結果'] = 'OK' if all(check_targets) else 'NG'

        result.append(report)
        print(f' ✔ 完了')
        ok_count += 1

    print(f' ✅ チェック完了: {ok_count} | チェック不可: {ng_count} | 合計: {len(target_files)}')

    print(f'作成チェック結果を出力します...')

    # 結果を出力
    result_df = pd.DataFrame(result)
    result_df = result_df.replace({True: 'OK', False: 'NG'})

    # チェック結果ファイルの作成(同名ファイルあれば連番付与)
    template_filepath = Path(config['ファイルパス']['案件アラートチェック結果FMT'])

    # 保存(すでにある場合はリネームして新規で作成)
    base_name = f"{YYYYMMDD}_案件アラート表不備チェック結果"
    save_filepath = serial_filepath(result_folder, base_name, '.xlsx')
    shutil.copy(template_filepath, save_filepath)

    
    # 全体の結果書き込み
    with xw.App(visible=False) as app:
        wb = xw.Book((save_filepath))
        ws = wb.sheets[0]  # 1番左のシートを指定

        
        ws.range("A:A").number_format = '@'
        ws.range('A10').options(index=False, header=False).value = result_df

        ws.range('C3').value = 'OK' if count_dict_len == target_files_len else 'NG'
        ws.range('C4').value = str(target_files_len)
        ws.range('C5').value = str(count_dict_len)
        ws.range('C6').value = 'OK' if all(result_df['トータルの不備結果'] == 'OK') else 'NG'
    
        # 保存
        wb.save()
        wb.close()

    print(f' >> {str(save_filepath)} を作成')


# ============================================
# 送付対象クライアント一覧の作成
# ============================================
def make_send_list():

    exist_folders([BASE_FOLDER])
    
    title = f'案件アラート表不備チェック結果.xlsxを選択してください'
    print(title)
    result_filepath = select_file(title=title, file_types=[('EXCELファイル', '*.xlsx')], initial_dir=BASE_FOLDER)
    result_folder = result_filepath.parent

    print(f' << {result_filepath} を取得')
        
    # 結果OKで絞り込み
    result_df = pd.read_excel(result_filepath, engine='openpyxl', skiprows=8, usecols='A:E', dtype=str)
    ok_df = result_df[result_df['トータルの不備結果'] == 'OK']
    ok_df = ok_df.iloc[:, :4]   # 4列目まで使用

    # データが0件（！エラー出力）
    if len(ok_df) == 0: raise ZeroDataError('OKが0件です。')

    # == フォーマット読込 ==
    template_filepath = config['ファイルパス']['送付対象クライアント一覧FMT']

    # 保存ファイル名(すでにある場合はリネームして新規で作成)
    base_name = f"{YYYYMMDD}_案件アラート送付対象クライアント一覧"
    save_filepath = serial_filepath(RESULT_FOLDER, base_name, '.xlsx')
    shutil.copy(template_filepath, save_filepath)

    # == 一括で書きこみ ==
    with xw.App(visible=False) as app:
        wb = xw.Book((save_filepath))
        ws = wb.sheets[0]  # 1番左のシートを指定
        
        ws.range("A:D").number_format = '@'
        ws.range("A2").options(index=False, header=False).value = ok_df
        
        # 保存
        wb.save()
        wb.close()

    print(f' >> {save_filepath} を作成')


# ============================================
# メイン処理
# ============================================
if __name__ == '__main__':

    try:
        # # ==== フォルダ作成 ====
        # make_folders()

        # # ==== 案件アラート作成 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 案件アラート作成開始')
        # create_sheets()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 案件アラート作成終了')

        # # ==== 案件アラートチェック ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] チェック開始')
        # check_sheets_at_creation()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] チェック終了')

        # # ==== 送付対象クライアント一覧出力 ====
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 送付対象クライアント一覧作成開始')
        # make_send_list()
        # print(f'[{datetime.datetime.now() :%Y-%m-%d %H:%M:%S}] 送付対象クライアント一覧作成終了')

        pass
    
    except KeyboardInterrupt:
        print('\n⛔ 処理を中断しました')
    except Exception as e:
        err_name = type(e).__name__
        err_msg = str(e)
        err_detail = traceback.format_exc()
        print(f"\n⚠️ エラーが発生しました \nエラー名: {err_name}\n詳細: {err_msg}\n{err_detail}")